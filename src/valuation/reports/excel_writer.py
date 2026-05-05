"""Generate Excel valuation workbook with multiple sheets.

Produces a structured Excel file similar to Damodaran's example spreadsheets,
with sheets for: Summary, Assumptions, DCF Model, Relative Valuation,
Sensitivity, Analyst Consensus, and Data Sources.

Color coding convention:
  - Light green (#E8F5E9): Facts/data from sources (revenue, net income, price)
  - Light blue (#E3F2FD): Assumptions (growth rate, WACC, beta, terminal growth)
  - No color: Calculations (EBIT(1-t), FCFF, PV, terminal value)
  - Light gray (#F5F5F5): Hardcoded values from formulas/heuristics

All formatting is deterministic Python -- no LLM calls.
"""

from __future__ import annotations

import pathlib
from datetime import date
from typing import Any

import pandas as pd
from openpyxl.comments import Comment
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from valuation.context import ValuationContext


# ---------------------------------------------------------------------------
# Colour / style constants
# ---------------------------------------------------------------------------

_DARK_BLUE = "1F3864"       # header row background
_LIGHT_GRAY = "D9D9D9"      # section-title background
_WHITE = "FFFFFF"
_GREEN = "00B050"            # positive upside
_RED = "FF0000"              # negative
_YELLOW_BG = "FFFF00"       # base-case cell fill in sensitivity

# NEW: semantic color fills
_FACT_FILL = PatternFill("solid", fgColor="E8F5E9")       # light green - facts/data
_ASSUMPTION_FILL = PatternFill("solid", fgColor="E3F2FD")  # light blue - assumptions
_HARDCODED_FILL = PatternFill("solid", fgColor="F5F5F5")   # light gray - hardcoded
# No fill for calculations (they use formulas)

_HEADER_FONT = Font(name="Calibri", bold=True, size=12, color=_WHITE)
_SECTION_FONT = Font(name="Calibri", bold=True, size=11, color="000000")
_NORMAL_FONT = Font(name="Calibri", size=10)
_BOLD_FONT = Font(name="Calibri", bold=True, size=10)

_HEADER_FILL = PatternFill("solid", fgColor=_DARK_BLUE)
_SECTION_FILL = PatternFill("solid", fgColor=_LIGHT_GRAY)
_WHITE_FILL = PatternFill("solid", fgColor=_WHITE)

_THIN = Side(style="thin")
_THIN_BORDER = Border(bottom=_THIN)

_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT = Alignment(horizontal="left", vertical="center")
_RIGHT = Alignment(horizontal="right", vertical="center")

# Number format strings
_FMT_INT = "#,##0"
_FMT_DEC = "#,##0.00"
_FMT_PCT = "0.0%"
_FMT_PCT2 = "0.00%"
_FMT_PRICE = "#,##0.00"


# ---------------------------------------------------------------------------
# Low-level helpers
# ---------------------------------------------------------------------------

def _safe(val: Any, fmt: str = ",.2f") -> Any:
    """Safe format: return value or empty string for None/NaN."""
    if val is None:
        return ""
    try:
        if isinstance(val, float) and val != val:  # NaN
            return ""
        return val
    except (TypeError, ValueError):
        return ""


def _ws(writer: pd.ExcelWriter, sheet_name: str):
    """Return the openpyxl worksheet for the given sheet name."""
    return writer.sheets[sheet_name]


def _hide_gridlines(ws) -> None:
    """Turn off gridlines on the worksheet."""
    ws.sheet_view.showGridLines = False


def _style_header_row(ws, row: int, n_cols: int) -> None:
    """Apply dark-blue header styling to an entire row."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _HEADER_FONT
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER if col > 1 else _LEFT


def _style_section_title(ws, row: int, n_cols: int) -> None:
    """Apply light-gray section-title styling to a row."""
    for col in range(1, n_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _SECTION_FONT
        cell.fill = _SECTION_FILL
        cell.alignment = _LEFT


def _autofit_columns(ws, min_label_width: int = 30, data_width: int = 16) -> None:
    """Set column A wider (for labels) and remaining columns to data_width."""
    ws.column_dimensions["A"].width = min_label_width
    for col_idx in range(2, ws.max_column + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = data_width


def _freeze_top_row(ws) -> None:
    ws.freeze_panes = "A2"


def _is_section_title(val: Any) -> bool:
    """Heuristic: a cell is a section title if it's an ALL-CAPS string."""
    if not isinstance(val, str):
        return False
    stripped = val.strip()
    return stripped == stripped.upper() and len(stripped) > 2 and stripped != ""


def _is_pct_label(label: str) -> bool:
    """Return True if the label suggests the value is a percentage."""
    label_lower = label.lower() if isinstance(label, str) else ""
    # Beta, D/E, S2C, PEG are ratios, NOT percentages
    non_pct = ("beta", "d/e", "sales to capital", "peg", "pe ", "pbv",
               "ev_ebitda", "ev_ebit", "ev_sales", "ev_invested", "current_pe",
               "trailing_pe", "forward_pe")
    if any(kw in label_lower for kw in non_pct):
        return False
    pct_keywords = ("rate", "growth", "yield", "wacc", "margin",
                    "upside", "downside", "completeness", "agreement",
                    "sensitivity", "score", "coverage", "premium", "discount",
                    "cost of equity", "cost of debt", "cost of capital",
                    "roe", "roic", "roc")
    return any(kw in label_lower for kw in pct_keywords)


def _apply_number_format(cell, label: str = "") -> None:
    """Apply number format based on value type and label hints."""
    val = cell.value
    if not isinstance(val, (int, float)):
        return
    if _is_pct_label(label) and abs(val) <= 3:
        # Likely stored as decimal fraction (e.g. 0.085 = 8.5%)
        cell.number_format = _FMT_PCT
    elif isinstance(val, float) and abs(val) < 1000 and abs(val) > 0:
        cell.number_format = _FMT_PRICE
    else:
        cell.number_format = _FMT_INT


def _apply_upside_color(cell) -> None:
    """Color a percentage upside/downside cell green or red."""
    val = cell.value
    if not isinstance(val, (int, float)):
        return
    if val > 0:
        cell.font = Font(name="Calibri", size=10, color=_GREEN)
    elif val < 0:
        cell.font = Font(name="Calibri", size=10, color=_RED)


def _add_comment(cell, text: str) -> None:
    """Add a comment/note to a cell."""
    cell.comment = Comment(text, "Valuation Agent")


def _safe_label(label: str) -> str:
    """Ensure a row label doesn't start with =, -, +, @ (which Excel interprets as formula)."""
    if isinstance(label, str) and label.strip() and label.strip()[0] in ("=", "-", "+", "@"):
        # Prefix with a space or rewrite
        stripped = label.strip()
        if stripped.startswith("= "):
            return stripped[2:]  # Remove "= " prefix
        elif stripped.startswith("- "):
            return "Less: " + stripped[2:]
        elif stripped.startswith("+ "):
            return "Add: " + stripped[2:]
        else:
            return "'" + label  # Escape with apostrophe
    return label


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def generate_excel(
    ctx: ValuationContext,
    ibes_data: dict | None = None,
    output_path: str | pathlib.Path | None = None,
) -> pathlib.Path:
    """Generate a multi-sheet Excel valuation workbook.

    Args:
        ctx: Completed ValuationContext with all outputs populated
        ibes_data: Optional dict with I/B/E/S analyst consensus data
        output_path: Where to save. If None, auto-generates path.

    Returns:
        Path to saved Excel file
    """
    if output_path is None:
        reports_dir = pathlib.Path(__file__).parent.parent.parent.parent / "reports"
        company_name = (ctx.company.name or ctx.company.ticker).replace("/", "-").replace("\\", "-")
        company_dir = reports_dir / company_name
        company_dir.mkdir(parents=True, exist_ok=True)
        ticker = ctx.company.ticker.replace(":", "-")
        output_path = company_dir / f"{date.today().isoformat()}_{ticker}.xlsx"

    output_path = pathlib.Path(output_path)

    # If file is locked (e.g., open in Excel), add a suffix
    if output_path.exists():
        try:
            output_path.unlink()
        except PermissionError:
            stem = output_path.stem
            output_path = output_path.with_name(f"{stem}_{date.today().strftime('%H%M%S')}.xlsx")

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        _write_summary(writer, ctx)
        _write_assumptions(writer, ctx)
        _write_dcf_model(writer, ctx)
        _write_relative_valuation(writer, ctx)
        _write_peer_comparison(writer, ctx)
        _write_sensitivity(writer, ctx)
        _write_analyst_consensus(writer, ctx, ibes_data)
        _write_data_sources(writer, ctx)
        _write_financials(writer, ctx)

    return output_path


# ---------------------------------------------------------------------------
# Sheet writers
# ---------------------------------------------------------------------------

def _write_summary(writer: pd.ExcelWriter, ctx: ValuationContext) -> None:
    """Sheet 1: Executive Summary -- key results at a glance."""
    a = ctx.assumptions
    price = ctx.financials.key_stats.get("price")

    # Collect all model values
    models = {}
    if ctx.outputs.dcf_fcff:
        models["DCF (FCFF)"] = ctx.outputs.dcf_fcff.get("equity_value_per_share")
    if ctx.outputs.dcf_fcfe:
        models["DDM"] = ctx.outputs.dcf_fcfe.get("value_per_share")
    if ctx.outputs.relative:
        models["Relative (PE)"] = ctx.outputs.relative.get("pe_value")
        models["Relative (EV/EBITDA)"] = ctx.outputs.relative.get("ev_ebitda_value")
        models["Relative (PBV)"] = ctx.outputs.relative.get("pbv_value")
        models["Relative (PS)"] = ctx.outputs.relative.get("ps_value")
        models["Relative (Composite)"] = ctx.outputs.relative.get("composite_value")
    if ctx.outputs.excess_returns:
        models["Excess Returns"] = ctx.outputs.excess_returns.get("value_per_share")

    values = [v for v in models.values() if v is not None and v > 0]

    rows = [
        ["VALUATION SUMMARY", ""],
        ["Date", date.today().isoformat()],
        ["", ""],
        ["Company", ctx.company.name or ctx.company.ticker],
        ["Ticker", ctx.company.ticker],
        ["Classification", ctx.company.classification],
        ["Damodaran Industry", ctx.company.damodaran_industry],
        ["Region", ctx.company.region],
        ["", ""],
        ["MARKET DATA", ""],
        ["Current Price", _safe(price)],
        ["Market Cap", _safe(ctx.financials.key_stats.get("market_cap"))],
        ["Shares Outstanding", _safe(ctx.financials.key_stats.get("shares_outstanding"))],
        ["", ""],
        ["KEY ASSUMPTIONS", ""],
        ["Risk-Free Rate", _safe(a.risk_free_rate)],
        ["Equity Risk Premium", _safe(a.erp)],
        ["Beta", _safe(a.beta)],
        ["Cost of Equity", _safe(a.cost_of_equity)],
        ["WACC", _safe(a.wacc)],
        ["Terminal Growth", _safe(a.terminal_growth)],
        ["", ""],
        ["VALUATION RESULTS", ""],
    ]

    has_upside = any(len(r) > 2 for r in rows)
    for model_name, val in models.items():
        if val is not None:
            upside = (val - price) / price if price and price > 0 else None
            rows.append([model_name, val, upside])
            has_upside = True
        else:
            rows.append([model_name, "N/A", ""])

    rows.append(["", ""])
    if values:
        rows.append(["Value Range (Low)", min(values)])
        rows.append(["Value Range (High)", max(values)])
        rows.append(["Mean", sum(values) / len(values)])
        if price and price > 0:
            mean_val = sum(values) / len(values)
            rows.append(["Implied Upside/Downside", (mean_val - price) / price])

    rows.append(["", ""])
    rows.append(["CONFIDENCE", ""])
    rows.append(["Data Completeness", _safe(ctx.confidence.data_completeness)])
    rows.append(["Model Agreement", _safe(ctx.confidence.model_agreement)])
    rows.append(["Assumption Sensitivity", _safe(ctx.confidence.assumption_sensitivity)])
    rows.append(["Composite Score", _safe(ctx.confidence.composite)])

    if ctx.confidence.flags:
        rows.append(["", ""])
        rows.append(["FLAGS", ""])
        for flag in ctx.confidence.flags:
            rows.append([flag, ""])

    # Company context
    profile = ctx.financials.key_stats.get("company_profile") or {}
    news = ctx.financials.key_stats.get("company_news") or []

    if profile.get("description"):
        rows.append(["", ""])
        rows.append(["COMPANY DESCRIPTION", ""])
        desc = profile["description"]
        for i in range(0, len(desc), 100):
            rows.append([desc[i:i + 100], ""])

    if news:
        rows.append(["", ""])
        rows.append(["RECENT NEWS", ""])
        for n in news[:5]:
            rows.append([n.get("title", ""), n.get("publisher", "")])

    transcript = ctx.financials.key_stats.get("earnings_transcript")
    if transcript:
        rows.append(["", ""])
        rows.append(["LATEST EARNINGS CALL", ""])
        rows.append([transcript.get("headline", ""), transcript.get("date", "")])
        # Add key excerpts (first 500 chars)
        excerpt = transcript.get("transcript_text", "")[:500]
        rows.append([excerpt, ""])

    cols = ["Item", "Value", "vs Market"] if has_upside else ["Item", "Value"]
    # Pad rows to match column count
    ncols = len(cols)
    for r in rows:
        while len(r) < ncols:
            r.append("")

    df = pd.DataFrame(rows, columns=cols)
    df.to_excel(writer, sheet_name="Summary", index=False)

    ws = _ws(writer, "Summary")
    _hide_gridlines(ws)

    # Company name as prominent title in A1 (merged)
    company_label = ctx.company.name or ctx.company.ticker
    ws.cell(row=1, column=1).value = f"Valuation Report \u2014 {company_label}"
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncols)
    ws.cell(row=1, column=1).font = Font(name="Calibri", bold=True, size=14, color=_WHITE)
    ws.cell(row=1, column=1).fill = _HEADER_FILL
    ws.cell(row=1, column=1).alignment = _CENTER

    # Walk through and apply styles + color coding
    pct_labels = {"risk-free rate", "equity risk premium", "beta", "cost of equity",
                  "wacc", "terminal growth", "data completeness", "model agreement",
                  "assumption sensitivity", "composite score", "implied upside/downside"}

    # Track which rows are facts vs assumptions for color coding
    fact_labels = {"current price", "market cap", "shares outstanding"}
    assumption_labels = {"risk-free rate", "equity risk premium", "beta",
                         "cost of equity", "wacc", "terminal growth"}

    for row_idx in range(2, ws.max_row + 1):
        label_cell = ws.cell(row=row_idx, column=1)
        label = label_cell.value or ""

        if _is_section_title(label):
            _style_section_title(ws, row_idx, ncols)
            continue

        label_cell.font = _BOLD_FONT if label and not isinstance(label, (int, float)) else _NORMAL_FONT
        label_cell.alignment = _LEFT

        # Value cell
        val_cell = ws.cell(row=row_idx, column=2)
        val = val_cell.value
        label_lower = str(label).lower()

        # Apply color coding
        if label_lower in fact_labels:
            val_cell.fill = _FACT_FILL
            _add_comment(val_cell, "Source: Yahoo Finance")
        elif label_lower in assumption_labels:
            val_cell.fill = _ASSUMPTION_FILL

        if isinstance(val, (int, float)):
            if label_lower in pct_labels and abs(val) <= 3:
                val_cell.number_format = _FMT_PCT
            elif label_lower in ("current price", "value range (low)", "value range (high)", "mean"):
                val_cell.number_format = _FMT_PRICE
            elif label_lower in ("market cap", "shares outstanding"):
                val_cell.number_format = _FMT_INT
            else:
                val_cell.number_format = _FMT_DEC
            val_cell.alignment = _RIGHT

        # vs Market column (upside/downside)
        if ncols >= 3:
            upside_cell = ws.cell(row=row_idx, column=3)
            if isinstance(upside_cell.value, (int, float)):
                upside_cell.number_format = _FMT_PCT
                _apply_upside_color(upside_cell)
                upside_cell.alignment = _RIGHT

    for col in range(1, ncols + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = Font(name="Calibri", bold=True, size=14, color=_WHITE)
        cell.fill = _HEADER_FILL
        cell.alignment = _CENTER

    _autofit_columns(ws)
    _freeze_top_row(ws)


def _write_assumptions(writer: pd.ExcelWriter, ctx: ValuationContext) -> None:
    """Sheet 2: All assumptions in a FIXED layout so DCF Model can cross-reference.

    Layout (rows are 1-indexed as they appear in the Excel file):
        Row 1:  Header row (Parameter / Value / Source)
        Row 2:  (blank)
        Row 3:  RISK PARAMETERS (section header)
        Row 4:  Risk-Free Rate          [B4]  blue
        Row 5:  Beta (levered)          [B5]  blue
        Row 6:  ERP                     [B6]  blue
        Row 7:  Lambda                  [B7]  blue
        Row 8:  CRP                     [B8]  blue
        Row 9:  Cost of Equity          [B9]  formula: =B4+B5*B6+B7*B8
        Row 10: Cost of Debt (pre-tax)  [B10] gray
        Row 11: Tax Rate (effective)    [B11] blue
        Row 12: Marginal Tax Rate       [B12] blue (for terminal)
        Row 13: Debt/(Debt+Equity)      [B13] green
        Row 14: WACC                    [B14] formula: =B9*(1-B13)+B10*(1-B11)*B13
        Row 15: (blank)
        Row 16: GROWTH & MARGINS (section header)
        Row 17: Revenue Growth (High)   [B17] blue
        Row 18: Terminal Growth          [B18] blue
        Row 19: Operating Margin         [B19] blue
        Row 20: Target Margin (Stable)   [B20] blue
        Row 21: Sales-to-Capital         [B21] blue
        Row 22: Stable ROC               [B22] blue
        Row 23: Convergence Year         [B23] blue
        Row 24: (blank)
        Row 25: COMPANY DATA (section header)
        Row 26: Base Revenue             [B26] green
        Row 27: Base EBIT                [B27] green
        Row 28: Cash                     [B28] green
        Row 29: Total Debt               [B29] green
        Row 30: Shares Outstanding       [B30] green
        Row 31: Current Price            [B31] green
    """
    a = ctx.assumptions

    # Extract company data from balance sheet / key_stats / dcf output
    d = ctx.outputs.dcf_fcff or {}
    stats = ctx.financials.key_stats

    # Base revenue and EBIT from dcf output or financials
    base_revenue = d.get("base_revenue")
    base_ebit = d.get("base_ebit")
    if base_revenue is None:
        yearly_rev = d.get("yearly_revenue", [])
        if yearly_rev and a.growth_rates:
            g0 = a.growth_rates[0]
            base_revenue = yearly_rev[0] / (1 + g0) if g0 != -1 else yearly_rev[0]
    if base_ebit is None and base_revenue:
        yearly_ebit = d.get("yearly_ebit", [])
        if yearly_ebit and yearly_rev:
            # Infer base margin from first-year margin applied to base revenue
            base_ebit = yearly_ebit[0] / (1 + (a.growth_rates[0] if a.growth_rates else 0))

    # Cash and debt from balance sheet
    cash = 0.0
    total_debt = 0.0
    if ctx.financials.balance_sheet is not None:
        bs = ctx.financials.balance_sheet
        # Handle both orientations
        if hasattr(bs, 'iloc') and len(bs) > 0:
            # Try to get latest year data
            try:
                if _looks_like_dates(bs.index):
                    latest = bs.iloc[0]
                elif _looks_like_dates(bs.columns):
                    latest = bs.iloc[:, 0]
                else:
                    latest = bs.iloc[0]
                cash = float(latest.get('Cash And Cash Equivalents', 0) or 0)
                total_debt = float(latest.get('Total Debt', 0) or 0)
            except (KeyError, TypeError, IndexError):
                pass

    shares = stats.get("shares_outstanding") or 0
    price = stats.get("price") or 0
    market_cap = stats.get("market_cap") or (price * shares if price and shares else 0)

    # Debt weight
    debt_weight = total_debt / (market_cap + total_debt) if (market_cap + total_debt) > 0 else 0

    # Lambda (default 1.0 unless we can find it from cost_of_equity decomposition)
    lambda_val = 1.0
    if (a.cost_of_equity is not None and a.risk_free_rate is not None
            and a.beta is not None and a.erp is not None
            and a.country_risk_premium and a.country_risk_premium > 0):
        # Ke = Rf + Beta*ERP + Lambda*CRP  =>  Lambda = (Ke - Rf - Beta*ERP) / CRP
        implied_lambda = (a.cost_of_equity - a.risk_free_rate - a.beta * a.erp) / a.country_risk_premium
        if 0 <= implied_lambda <= 2:
            lambda_val = round(implied_lambda, 4)

    # Operating margin (current)
    current_margin = 0.0
    if base_revenue and base_ebit and base_revenue > 0:
        current_margin = base_ebit / base_revenue

    # Target margin and S2C from dcf output schedules
    yearly_ebit_list = d.get("yearly_ebit", [])
    yearly_rev_list = d.get("yearly_revenue", [])
    target_margin = current_margin
    if yearly_ebit_list and yearly_rev_list and yearly_rev_list[-1]:
        target_margin = yearly_ebit_list[-1] / yearly_rev_list[-1]

    # Sales-to-capital from reinvestment data
    s2c = 2.0  # default
    yearly_reinv = d.get("yearly_reinvestment", [])
    if yearly_rev_list and yearly_reinv:
        for t in range(len(yearly_reinv)):
            prev_rev = base_revenue if t == 0 else yearly_rev_list[t - 1]
            curr_rev = yearly_rev_list[t] if t < len(yearly_rev_list) else 0
            reinv = yearly_reinv[t]
            if reinv and reinv != 0:
                s2c = (curr_rev - prev_rev) / reinv if prev_rev else s2c
                break  # use first year's implied S2C

    # Stable ROC
    stable_roc = 0.20  # default
    yearly_roic = d.get("yearly_roic", [])
    if yearly_roic and yearly_roic[-1] and yearly_roic[-1] > 0:
        stable_roc = yearly_roic[-1]

    # Revenue growth (high)
    high_growth = a.growth_rates[0] if a.growth_rates else 0.0

    # Marginal tax rate (for terminal year)
    marginal_tax = 0.25  # Damodaran convention for terminal
    # If we have tax schedule data in the dcf output, use last year
    yearly_ebit_at = d.get("yearly_ebit_at", [])
    if yearly_ebit_list and yearly_ebit_at and yearly_ebit_list[-1] and yearly_ebit_list[-1] != 0:
        terminal_tax_implied = 1 - yearly_ebit_at[-1] / yearly_ebit_list[-1]
        if 0 < terminal_tax_implied < 0.5:
            marginal_tax = terminal_tax_implied

    # Convergence year (how many years margin converges)
    convergence_year = 5  # default

    # --- Write the fixed-layout sheet using openpyxl directly ---
    empty_df = pd.DataFrame([[""] * 3])
    empty_df.to_excel(writer, sheet_name="Assumptions", index=False, header=False)
    ws = _ws(writer, "Assumptions")
    _hide_gridlines(ws)

    # Row 1: Header
    ws.cell(row=1, column=1, value="Parameter")
    ws.cell(row=1, column=2, value="Value")
    ws.cell(row=1, column=3, value="Source / Rationale")
    _style_header_row(ws, 1, 3)

    # Row 2: blank

    # Row 3: RISK PARAMETERS section header
    ws.cell(row=3, column=1, value="RISK PARAMETERS")
    _style_section_title(ws, 3, 3)

    # Row 4: Risk-Free Rate [B4] blue
    ws.cell(row=4, column=1, value="Risk-Free Rate").font = _NORMAL_FONT
    cell = ws.cell(row=4, column=2, value=a.risk_free_rate or 0)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=4, column=3, value="Damodaran / Govt bond yield").font = _NORMAL_FONT

    # Row 5: Beta (levered) [B5] blue
    ws.cell(row=5, column=1, value="Beta (levered)").font = _NORMAL_FONT
    cell = ws.cell(row=5, column=2, value=a.beta or 0)
    cell.number_format = _FMT_DEC
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=5, column=3, value="Industry bottom-up, re-levered").font = _NORMAL_FONT

    # Row 6: ERP [B6] blue
    ws.cell(row=6, column=1, value="Equity Risk Premium").font = _NORMAL_FONT
    cell = ws.cell(row=6, column=2, value=a.erp or 0)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=6, column=3, value="Damodaran implied ERP").font = _NORMAL_FONT

    # Row 7: Lambda [B7] blue
    ws.cell(row=7, column=1, value="Lambda (country exposure)").font = _NORMAL_FONT
    cell = ws.cell(row=7, column=2, value=lambda_val)
    cell.number_format = _FMT_DEC
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=7, column=3, value="Firm-specific country risk exposure (0-1)").font = _NORMAL_FONT

    # Row 8: CRP [B8] blue
    ws.cell(row=8, column=1, value="Country Risk Premium").font = _NORMAL_FONT
    cell = ws.cell(row=8, column=2, value=a.country_risk_premium or 0)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=8, column=3, value="Damodaran ctryprem").font = _NORMAL_FONT

    # Row 9: Cost of Equity [B9] formula
    ws.cell(row=9, column=1, value="Cost of Equity").font = _NORMAL_FONT
    cell = ws.cell(row=9, column=2, value="=B4+B5*B6+B7*B8")
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    ws.cell(row=9, column=3, value="Formula: Rf + Beta*ERP + Lambda*CRP").font = _NORMAL_FONT

    # Row 10: Cost of Debt (pre-tax) [B10] gray
    ws.cell(row=10, column=1, value="Cost of Debt (pre-tax)").font = _NORMAL_FONT
    cell = ws.cell(row=10, column=2, value=a.cost_of_debt or 0)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _HARDCODED_FILL
    ws.cell(row=10, column=3, value="Synthetic rating spread + Rf").font = _NORMAL_FONT

    # Row 11: Tax Rate (effective) [B11] blue
    ws.cell(row=11, column=1, value="Tax Rate (effective)").font = _NORMAL_FONT
    cell = ws.cell(row=11, column=2, value=a.tax_rate or 0)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=11, column=3, value="Effective from financials").font = _NORMAL_FONT

    # Row 12: Marginal Tax Rate [B12] blue
    ws.cell(row=12, column=1, value="Marginal Tax Rate").font = _NORMAL_FONT
    cell = ws.cell(row=12, column=2, value=marginal_tax)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=12, column=3, value="For terminal year (Damodaran convention)").font = _NORMAL_FONT

    # Row 13: Debt/(Debt+Equity) [B13] green
    ws.cell(row=13, column=1, value="Debt/(Debt+Equity)").font = _NORMAL_FONT
    cell = ws.cell(row=13, column=2, value=debt_weight)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL
    ws.cell(row=13, column=3, value="From balance sheet & market cap").font = _NORMAL_FONT

    # Row 14: WACC [B14] formula
    ws.cell(row=14, column=1, value="WACC").font = _NORMAL_FONT
    cell = ws.cell(row=14, column=2, value="=B9*(1-B13)+B10*(1-B11)*B13")
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    ws.cell(row=14, column=3, value="Formula: Ke*(1-Wd) + Kd*(1-t)*Wd").font = _NORMAL_FONT

    # Row 15: blank

    # Row 16: GROWTH & MARGINS section header
    ws.cell(row=16, column=1, value="GROWTH & MARGINS")
    _style_section_title(ws, 16, 3)

    # Row 17: Revenue Growth (High) [B17] blue
    ws.cell(row=17, column=1, value="Revenue Growth (High)").font = _NORMAL_FONT
    cell = ws.cell(row=17, column=2, value=high_growth)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=17, column=3, value="Fundamental or historical CAGR").font = _NORMAL_FONT

    # Row 18: Terminal Growth [B18] blue
    ws.cell(row=18, column=1, value="Terminal Growth").font = _NORMAL_FONT
    cell = ws.cell(row=18, column=2, value=a.terminal_growth or 0)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=18, column=3, value="Nominal GDP growth").font = _NORMAL_FONT

    # Row 19: Operating Margin [B19] blue
    ws.cell(row=19, column=1, value="Operating Margin (Current)").font = _NORMAL_FONT
    cell = ws.cell(row=19, column=2, value=current_margin)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=19, column=3, value="EBIT / Revenue (current)").font = _NORMAL_FONT

    # Row 20: Target Margin (Stable) [B20] blue
    ws.cell(row=20, column=1, value="Target Margin (Stable)").font = _NORMAL_FONT
    cell = ws.cell(row=20, column=2, value=target_margin)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=20, column=3, value="Margin at convergence").font = _NORMAL_FONT

    # Row 21: Sales-to-Capital [B21] blue
    ws.cell(row=21, column=1, value="Sales-to-Capital Ratio").font = _NORMAL_FONT
    cell = ws.cell(row=21, column=2, value=s2c)
    cell.number_format = _FMT_DEC
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=21, column=3, value="Revenue / Invested Capital").font = _NORMAL_FONT

    # Row 22: Stable ROC [B22] blue
    ws.cell(row=22, column=1, value="Stable ROC").font = _NORMAL_FONT
    cell = ws.cell(row=22, column=2, value=stable_roc)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=22, column=3, value="Return on capital in stable state").font = _NORMAL_FONT

    # Row 23: Convergence Year [B23] blue
    ws.cell(row=23, column=1, value="Convergence Year").font = _NORMAL_FONT
    cell = ws.cell(row=23, column=2, value=convergence_year)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    ws.cell(row=23, column=3, value="Year by which margins reach target").font = _NORMAL_FONT

    # Row 24: blank

    # Row 25: COMPANY DATA section header
    ws.cell(row=25, column=1, value="COMPANY DATA")
    _style_section_title(ws, 25, 3)

    # Row 26: Base Revenue [B26] green
    ws.cell(row=26, column=1, value="Base Revenue").font = _NORMAL_FONT
    cell = ws.cell(row=26, column=2, value=base_revenue or 0)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL
    ws.cell(row=26, column=3, value="TTM or latest annual (Yahoo Finance)").font = _NORMAL_FONT

    # Row 27: Base EBIT [B27] green
    ws.cell(row=27, column=1, value="Base EBIT").font = _NORMAL_FONT
    cell = ws.cell(row=27, column=2, value=base_ebit or 0)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL
    ws.cell(row=27, column=3, value="Operating Income (Yahoo Finance)").font = _NORMAL_FONT

    # Row 28: Cash [B28] green
    ws.cell(row=28, column=1, value="Cash & Equivalents").font = _NORMAL_FONT
    cell = ws.cell(row=28, column=2, value=cash)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL
    ws.cell(row=28, column=3, value="Balance Sheet (Yahoo Finance)").font = _NORMAL_FONT

    # Row 29: Total Debt [B29] green
    ws.cell(row=29, column=1, value="Total Debt").font = _NORMAL_FONT
    cell = ws.cell(row=29, column=2, value=total_debt)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL
    ws.cell(row=29, column=3, value="Balance Sheet (Yahoo Finance)").font = _NORMAL_FONT

    # Row 30: Shares Outstanding [B30] green
    ws.cell(row=30, column=1, value="Shares Outstanding").font = _NORMAL_FONT
    cell = ws.cell(row=30, column=2, value=shares)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL
    ws.cell(row=30, column=3, value="Yahoo Finance").font = _NORMAL_FONT

    # Row 31: Current Price [B31] green
    ws.cell(row=31, column=1, value="Current Price").font = _NORMAL_FONT
    cell = ws.cell(row=31, column=2, value=price)
    cell.number_format = _FMT_PRICE
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL
    ws.cell(row=31, column=3, value="Yahoo Finance (market price)").font = _NORMAL_FONT

    # --- Additional info rows below the fixed layout (row 33+) ---
    row = 33

    # Overrides
    if a.overrides:
        ws.cell(row=row, column=1, value="ANALYST OVERRIDES")
        _style_section_title(ws, row, 3)
        row += 1
        ws.cell(row=row, column=1, value="Parameter").font = _BOLD_FONT
        ws.cell(row=row, column=2, value="New Value").font = _BOLD_FONT
        ws.cell(row=row, column=3, value="Reason").font = _BOLD_FONT
        row += 1
        for param, info in a.overrides.items():
            ws.cell(row=row, column=1, value=param).font = _NORMAL_FONT
            ws.cell(row=row, column=2, value=info.get("new")).alignment = _RIGHT
            ws.cell(row=row, column=3, value=info.get("reason", "")).font = _NORMAL_FONT
            row += 1
        row += 1

    # Industry benchmarks
    bm = ctx.benchmarks
    ws.cell(row=row, column=1, value="INDUSTRY BENCHMARKS")
    _style_section_title(ws, row, 3)
    row += 1
    for label, val, src in [
        ("Industry Beta", bm.industry_beta, "Damodaran betas"),
        ("Industry Unlevered Beta", bm.industry_unlevered_beta, "Damodaran betas"),
        ("Industry D/E", bm.industry_de_ratio, "Damodaran betas"),
        ("Industry WACC", bm.industry_wacc, "Damodaran wacc"),
    ]:
        ws.cell(row=row, column=1, value=label).font = _NORMAL_FONT
        cell = ws.cell(row=row, column=2, value=_safe(val))
        if isinstance(val, (int, float)):
            cell.number_format = _FMT_DEC
            cell.alignment = _RIGHT
            cell.fill = _FACT_FILL
        ws.cell(row=row, column=3, value=src).font = _NORMAL_FONT
        row += 1

    if bm.industry_multiples:
        row += 1
        ws.cell(row=row, column=1, value="INDUSTRY MULTIPLES")
        _style_section_title(ws, row, 3)
        row += 1
        for k, v in bm.industry_multiples.items():
            ws.cell(row=row, column=1, value=k).font = _NORMAL_FONT
            cell = ws.cell(row=row, column=2, value=_safe(v))
            if isinstance(v, (int, float)):
                cell.number_format = _FMT_DEC
                cell.alignment = _RIGHT
            ws.cell(row=row, column=3, value="Damodaran").font = _NORMAL_FONT
            row += 1

    if bm.industry_margins:
        row += 1
        ws.cell(row=row, column=1, value="INDUSTRY MARGINS")
        _style_section_title(ws, row, 3)
        row += 1
        for k, v in bm.industry_margins.items():
            ws.cell(row=row, column=1, value=k).font = _NORMAL_FONT
            cell = ws.cell(row=row, column=2, value=_safe(v))
            if isinstance(v, (int, float)):
                cell.number_format = _FMT_PCT if abs(v) <= 1 else _FMT_DEC
                cell.alignment = _RIGHT
            ws.cell(row=row, column=3, value="Damodaran").font = _NORMAL_FONT
            row += 1

    # Column widths
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 16
    ws.column_dimensions["C"].width = 40
    ws.freeze_panes = "A2"


def _write_dcf_model(writer: pd.ExcelWriter, ctx: ValuationContext) -> None:
    """Sheet 3: DCF year-by-year projections -- formula-based Damodaran-style.

    Instead of hardcoding computed values, writes Excel formulas so the
    spreadsheet is interactive. Assumptions are written to named cells and
    projections reference them via formulas.
    """
    if ctx.outputs.dcf_fcff:
        d = ctx.outputs.dcf_fcff
        a = ctx.assumptions

        yearly_revenue = d.get("yearly_revenue", [])
        yearly_ebit = d.get("yearly_ebit", [])
        yearly_ebit_at = d.get("yearly_ebit_at", [])
        yearly_reinvestment = d.get("yearly_reinvestment", [])
        yearly_fcff = d.get("yearly_fcff", [])
        yearly_pv = d.get("yearly_pv", [])
        n = len(yearly_fcff)

        # Detect v2 (revenue-based) vs v1 (ebit-based)
        is_v2 = bool(yearly_revenue)

        if is_v2:
            _write_dcf_model_v2_formulas(writer, ctx, d, n)
        else:
            _write_dcf_model_v1(writer, ctx, d, n)

    elif ctx.outputs.dcf_fcfe:
        d = ctx.outputs.dcf_fcfe
        n = len(d.get("yearly_eps", []))
        rows = {
            "Year": list(range(1, n + 1)),
            "EPS": d.get("yearly_eps", []),
            "DPS": d.get("yearly_dps", []),
            "PV of DPS": d.get("yearly_pv", []),
        }
        df = pd.DataFrame(rows)
        summary = pd.DataFrame([
            {"Year": "", "EPS": "", "DPS": "", "PV of DPS": ""},
            {"Year": "PV of Dividends", "EPS": "", "DPS": "", "PV of DPS": d.get("pv_dividends")},
            {"Year": "Terminal Price", "EPS": "", "DPS": d.get("terminal_price"), "PV of DPS": d.get("pv_terminal")},
            {"Year": "Value per Share", "EPS": "", "DPS": "", "PV of DPS": d.get("value_per_share")},
        ])
        df = pd.concat([df, summary], ignore_index=True)
        df.to_excel(writer, sheet_name="DCF Model (DDM)", index=False)

        ws = _ws(writer, "DCF Model (DDM)")
        _hide_gridlines(ws)
        _style_header_row(ws, 1, 4)

        for row_idx in range(2, ws.max_row + 1):
            year_cell = ws.cell(row=row_idx, column=1)
            year_val = year_cell.value
            if isinstance(year_val, str) and year_val.strip():
                year_cell.font = _BOLD_FONT
            for col_idx in range(2, 5):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    cell.number_format = _FMT_PRICE
                    cell.alignment = _RIGHT

        _autofit_columns(ws, min_label_width=20)
        _freeze_top_row(ws)


def _write_dcf_model_v2_formulas(
    writer: pd.ExcelWriter, ctx: ValuationContext, d: dict, n: int
) -> None:
    """Write the DCF Model sheet using Excel formulas referencing the Assumptions sheet.

    All key inputs (WACC, growth, margins, cash, debt, shares) are pulled from
    the Assumptions sheet at known cell positions, making the workbook fully dynamic.

    Layout:
      Col A: Labels
      Col B: Base year
      Col C..C+n-1: Year 1..n
      Col C+n: Terminal year

    Assumptions sheet references:
      B4=Rf, B5=Beta, B6=ERP, B7=Lambda, B8=CRP, B9=Ke, B10=Kd,
      B11=TaxRate, B12=MarginalTax, B13=DebtWeight, B14=WACC,
      B17=HighGrowth, B18=TerminalGrowth, B19=CurrentMargin,
      B20=TargetMargin, B21=S2C, B22=StableROC, B23=ConvergenceYear,
      B26=BaseRevenue, B27=BaseEBIT, B28=Cash, B29=Debt,
      B30=Shares, B31=Price
    """
    a = ctx.assumptions
    yearly_revenue = d.get("yearly_revenue", [])
    yearly_ebit = d.get("yearly_ebit", [])
    yearly_ebit_at = d.get("yearly_ebit_at", [])
    yearly_reinvestment = d.get("yearly_reinvestment", [])
    yearly_fcff = d.get("yearly_fcff", [])
    yearly_pv = d.get("yearly_pv", [])

    base_revenue = d.get("base_revenue")
    base_ebit = d.get("base_ebit")

    if base_revenue is None and yearly_revenue and a.growth_rates:
        g0 = a.growth_rates[0]
        base_revenue = yearly_revenue[0] / (1 + g0) if g0 != -1 else yearly_revenue[0]

    base_margin = (base_ebit / base_revenue) if (base_ebit is not None and base_revenue) else 0

    # Compute per-year WACC from cumulated discount factors (for hardcoded schedule)
    cum_discount: list = []
    for t in range(n):
        fcff_t = yearly_fcff[t] if t < len(yearly_fcff) else 0
        pv_t = yearly_pv[t] if t < len(yearly_pv) else 0
        if fcff_t and fcff_t != 0 and pv_t and pv_t != 0:
            cum_discount.append(abs(fcff_t / pv_t))
        else:
            cum_discount.append(1.0)

    wacc_y: list = []
    for t in range(n):
        cd = cum_discount[t]
        if t == 0:
            wacc_y.append(cd - 1)
        else:
            prev_cd = cum_discount[t - 1]
            if prev_cd and prev_cd != 0:
                wacc_y.append(cd / prev_cd - 1)
            else:
                wacc_y.append(a.wacc if a.wacc is not None else 0)

    # Terminal FCFF (from engine, needed as fallback)
    terminal_fcff = d.get("terminal_fcff")

    # --- Write to worksheet using openpyxl directly ---
    empty_df = pd.DataFrame([[""] * (n + 3)])
    empty_df.to_excel(writer, sheet_name="DCF Model", index=False, header=False)
    ws = _ws(writer, "DCF Model")
    _hide_gridlines(ws)

    # Column layout: A=labels, B=Base, C=Year1, D=Year2, ... C+n-1=YearN, C+n=Terminal
    max_cols = n + 3  # A + Base + n years + Terminal
    term_col = n + 3  # terminal column number (1-indexed)

    def col_let(c: int) -> str:
        return get_column_letter(c)

    # --- Row 1: Header row ---
    row = 1
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="Base")
    for t in range(n):
        ws.cell(row=row, column=t + 3, value=f"Year {t+1}")
    ws.cell(row=row, column=term_col, value="Terminal")
    _style_header_row(ws, row, max_cols)

    # --- Row 2: Section: REVENUE PROJECTIONS ---
    row = 2
    ws.cell(row=row, column=1, value="REVENUE PROJECTIONS")
    _style_section_title(ws, row, max_cols)

    # --- Row 3: Revenue Growth Rate ---
    # Years 1-5: =Assumptions!$B$17 (high growth)
    # Years 6-n: linearly interpolate from high growth to terminal growth
    # Terminal col: =Assumptions!$B$18
    row = 3
    ws.cell(row=row, column=1, value="Revenue Growth Rate")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    ws.cell(row=row, column=2, value="")  # no growth for base

    # Determine the interpolation boundary (first 5 years = constant high growth)
    n_constant = min(5, n)
    for t in range(n):
        c = t + 3
        if t < n_constant:
            # First n_constant years: reference high growth from Assumptions
            formula = "=Assumptions!$B$17"
        else:
            # Linear interpolation from high growth to terminal growth
            # Formula: =Assumptions!$B$17 - (Assumptions!$B$17 - Assumptions!$B$18) * (year - n_constant) / (n - n_constant)
            steps = n - n_constant
            step_num = t - n_constant + 1
            formula = f"=Assumptions!$B$17-(Assumptions!$B$17-Assumptions!$B$18)*{step_num}/{steps}"
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = _FMT_PCT
        cell.alignment = _RIGHT
        cell.fill = _ASSUMPTION_FILL

    # Terminal growth
    cell = ws.cell(row=row, column=term_col, value="=Assumptions!$B$18")
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL

    # --- Row 4: Revenues ---
    # Base: =Assumptions!$B$26
    # Year t: =prev_col_4 * (1 + this_col_3)
    row = 4
    ws.cell(row=row, column=1, value="Revenues")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    cell = ws.cell(row=row, column=2, value="=Assumptions!$B$26")
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL
    for t in range(n):
        c = t + 3
        prev_cl = col_let(c - 1)
        growth_cl = col_let(c)
        formula = f"={prev_cl}4*(1+{growth_cl}3)"
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = _FMT_INT
        cell.alignment = _RIGHT

    # --- Row 5: Operating Margin ---
    # Margin convergence: =Assumptions!$B$19 + (Assumptions!$B$20 - Assumptions!$B$19) * MIN(year/Assumptions!$B$23, 1)
    row = 5
    ws.cell(row=row, column=1, value="Operating Margin")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    # Base margin: =Assumptions!$B$27 / Assumptions!$B$26 (EBIT/Revenue)
    cell = ws.cell(row=row, column=2, value="=IF(Assumptions!$B$26=0,0,Assumptions!$B$27/Assumptions!$B$26)")
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL
    for t in range(n):
        c = t + 3
        year_num = t + 1
        formula = f"=Assumptions!$B$19+(Assumptions!$B$20-Assumptions!$B$19)*MIN({year_num}/Assumptions!$B$23,1)"
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = _FMT_PCT
        cell.alignment = _RIGHT
        cell.fill = _ASSUMPTION_FILL
    # Terminal margin = target margin
    cell = ws.cell(row=row, column=term_col, value="=Assumptions!$B$20")
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL

    # --- Row 6: Operating Income (EBIT) = Revenue * Margin ---
    row = 6
    ws.cell(row=row, column=1, value="Operating Income (EBIT)")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    cell = ws.cell(row=row, column=2, value="=Assumptions!$B$27")
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL
    for t in range(n):
        c = t + 3
        cl = col_let(c)
        formula = f"={cl}4*{cl}5"
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = _FMT_INT
        cell.alignment = _RIGHT

    # --- Row 7: blank ---

    # --- Row 8: Section: TAX & REINVESTMENT ---
    row = 8
    ws.cell(row=row, column=1, value="TAX & REINVESTMENT")
    _style_section_title(ws, row, max_cols)

    # --- Row 9: Tax Rate ---
    # Years 1-5: =Assumptions!$B$11 (effective)
    # Years 6-n: ramp to marginal =Assumptions!$B$12
    # Terminal: =Assumptions!$B$12
    row = 9
    ws.cell(row=row, column=1, value="Tax Rate")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    cell = ws.cell(row=row, column=2, value="=Assumptions!$B$11")
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL
    for t in range(n):
        c = t + 3
        if t < n_constant:
            formula = "=Assumptions!$B$11"
        else:
            steps = n - n_constant
            step_num = t - n_constant + 1
            formula = f"=Assumptions!$B$11+(Assumptions!$B$12-Assumptions!$B$11)*{step_num}/{steps}"
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = _FMT_PCT
        cell.alignment = _RIGHT
        cell.fill = _ASSUMPTION_FILL
    cell = ws.cell(row=row, column=term_col, value="=Assumptions!$B$12")
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _ASSUMPTION_FILL

    # --- Row 10: EBIT(1-t) = EBIT * (1 - TaxRate) ---
    row = 10
    ws.cell(row=row, column=1, value="EBIT(1-t)")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    cell = ws.cell(row=row, column=2, value="=B6*(1-B9)")
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    for t in range(n):
        c = t + 3
        cl = col_let(c)
        formula = f"={cl}6*(1-{cl}9)"
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = _FMT_INT
        cell.alignment = _RIGHT

    # --- Row 11: Sales to Capital Ratio ---
    # All years reference Assumptions!$B$21
    row = 11
    ws.cell(row=row, column=1, value="Sales to Capital Ratio")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    for t in range(n):
        c = t + 3
        cell = ws.cell(row=row, column=c, value="=Assumptions!$B$21")
        cell.number_format = _FMT_DEC
        cell.alignment = _RIGHT
        cell.fill = _ASSUMPTION_FILL

    # --- Row 12: Reinvestment = (Revenue - PrevRevenue) / S2C ---
    row = 12
    ws.cell(row=row, column=1, value="Reinvestment")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    for t in range(n):
        c = t + 3
        cl = col_let(c)
        prev_cl = col_let(c - 1)
        formula = f"=({cl}4-{prev_cl}4)/{cl}11"
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = _FMT_INT
        cell.alignment = _RIGHT

    # --- Row 13: FCFF = EBIT(1-t) - Reinvestment ---
    row = 13
    ws.cell(row=row, column=1, value="Free Cash Flow to Firm")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    for t in range(n):
        c = t + 3
        cl = col_let(c)
        formula = f"={cl}10-{cl}12"
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = _FMT_INT
        cell.alignment = _RIGHT
    # Terminal FCFF: =last_year_EBIT_at * (1 + g) * (1 - g/ROC)
    term_cl = col_let(term_col)
    last_year_cl = col_let(term_col - 1)
    # Terminal FCFF = EBIT(1-t)_last_year * (1+g) * (1 - g/ROC)
    formula = f"={last_year_cl}10*(1+Assumptions!$B$18)*(1-Assumptions!$B$18/Assumptions!$B$22)"
    cell = ws.cell(row=row, column=term_col, value=formula)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT

    # --- Row 14: blank ---

    # --- Row 15: Section: DISCOUNT RATES ---
    row = 15
    ws.cell(row=row, column=1, value="DISCOUNT RATES")
    _style_section_title(ws, row, max_cols)

    # --- Row 16: Cost of Capital (WACC) ---
    # Years 1-5: =Assumptions!$B$14
    # Years 6-n: hardcoded schedule values (transition to stable)
    # Terminal: last value of schedule
    row = 16
    ws.cell(row=row, column=1, value="Cost of Capital (WACC)")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    for t in range(n):
        c = t + 3
        if t < n_constant:
            formula = "=Assumptions!$B$14"
        else:
            # Linear ramp: WACC transitions toward a stable WACC
            # Stable WACC approx = Rf + 4.5% + CRP (Damodaran convention)
            # Use formula: =Assumptions!$B$14 - (Assumptions!$B$14 - (Assumptions!$B$4+0.045+Assumptions!$B$8)) * step/steps
            steps = n - n_constant
            step_num = t - n_constant + 1
            # Simpler: hardcode from computed schedule since stable WACC formula is complex
            cell_val = wacc_y[t] if t < len(wacc_y) else (a.wacc or 0)
            cell = ws.cell(row=row, column=c, value=cell_val)
            cell.number_format = _FMT_PCT
            cell.alignment = _RIGHT
            cell.fill = _HARDCODED_FILL
            _add_comment(cell, "WACC transition (from schedule)")
            continue
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = _FMT_PCT
        cell.alignment = _RIGHT
        cell.fill = _ASSUMPTION_FILL
    # Terminal WACC (use last wacc_y value or hardcode)
    terminal_wacc = wacc_y[-1] if wacc_y else (a.wacc or 0)
    cell = ws.cell(row=row, column=term_col, value=terminal_wacc)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT
    cell.fill = _HARDCODED_FILL
    _add_comment(cell, "Stable-state WACC")

    # --- Row 17: Cumulated Discount Factor = prev * (1+WACC) ---
    row = 17
    ws.cell(row=row, column=1, value="Cumulated Discount Factor")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    for t in range(n):
        c = t + 3
        cl = col_let(c)
        if t == 0:
            formula = f"=1+{cl}16"
        else:
            prev_cl = col_let(c - 1)
            formula = f"={prev_cl}17*(1+{cl}16)"
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = "#,##0.0000"
        cell.alignment = _RIGHT

    # --- Row 18: PV(FCFF) = FCFF / CumulatedDiscount ---
    row = 18
    ws.cell(row=row, column=1, value="PV(FCFF)")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    for t in range(n):
        c = t + 3
        cl = col_let(c)
        formula = f"={cl}13/{cl}17"
        cell = ws.cell(row=row, column=c, value=formula)
        cell.number_format = _FMT_INT
        cell.alignment = _RIGHT

    # --- Row 19: blank ---

    # --- Row 20: Section: TERMINAL VALUE ---
    row = 20
    ws.cell(row=row, column=1, value="TERMINAL VALUE")
    _style_section_title(ws, row, max_cols)

    # --- Row 21: Terminal Value = Terminal FCFF / (Terminal WACC - Terminal Growth) ---
    row = 21
    ws.cell(row=row, column=1, value="Terminal Value")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    formula = f"={term_cl}13/({term_cl}16-{term_cl}3)"
    cell = ws.cell(row=row, column=term_col, value=formula)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT

    # --- Row 22: PV(Terminal Value) = TV / CumDiscount_last_year ---
    row = 22
    ws.cell(row=row, column=1, value="PV(Terminal Value)")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    formula = f"={term_cl}21/{last_year_cl}17"
    cell = ws.cell(row=row, column=term_col, value=formula)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT

    # --- Row 23: blank ---

    # --- Row 24: Section: EQUITY VALUE BRIDGE ---
    row = 24
    ws.cell(row=row, column=1, value="EQUITY VALUE BRIDGE")
    _style_section_title(ws, row, max_cols)

    non_op = d.get("non_operating_assets", 0) or 0

    # --- Row 25: PV of Operating Cash Flows = SUM(PV row) ---
    row = 25
    ws.cell(row=row, column=1, value="PV of Cash Flows (High Growth)")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    first_year_cl = col_let(3)
    last_data_cl = col_let(n + 2)
    formula = f"=SUM({first_year_cl}18:{last_data_cl}18)"
    cell = ws.cell(row=row, column=2, value=formula)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT

    # --- Row 26: PV of Terminal Value ---
    row = 26
    ws.cell(row=row, column=1, value="PV of Terminal Value")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    formula = f"={term_cl}22"
    cell = ws.cell(row=row, column=2, value=formula)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT

    # --- Row 27: Value of Operating Assets ---
    row = 27
    ws.cell(row=row, column=1, value="Value of Operating Assets")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    formula = "=B25+B26"
    cell = ws.cell(row=row, column=2, value=formula)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT

    # --- Row 28: + Cash [references Assumptions!B28] ---
    row = 28
    ws.cell(row=row, column=1, value="Add: Cash & Marketable Securities")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    cell = ws.cell(row=row, column=2, value="=Assumptions!$B$28")
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL

    # --- Row 29: - Debt [references Assumptions!B29] ---
    row = 29
    ws.cell(row=row, column=1, value="Less: Debt")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    cell = ws.cell(row=row, column=2, value="=Assumptions!$B$29")
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL

    # --- Row 30: + Non-operating Assets ---
    row = 30
    ws.cell(row=row, column=1, value="Add: Non-operating Assets")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    cell = ws.cell(row=row, column=2, value=non_op)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    if non_op:
        cell.fill = _HARDCODED_FILL

    # --- Row 31: = Value of Equity ---
    row = 31
    ws.cell(row=row, column=1, value="Value of Equity")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    formula = "=B27+B28-B29+B30"
    cell = ws.cell(row=row, column=2, value=formula)
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT

    # --- Row 32: / Shares Outstanding [references Assumptions!B30] ---
    row = 32
    ws.cell(row=row, column=1, value="Shares Outstanding")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    cell = ws.cell(row=row, column=2, value="=Assumptions!$B$30")
    cell.number_format = _FMT_INT
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL

    # --- Row 33: = Value per Share ---
    row = 33
    ws.cell(row=row, column=1, value="Value per Share")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    formula = "=B31/B32"
    cell = ws.cell(row=row, column=2, value=formula)
    cell.number_format = _FMT_PRICE
    cell.alignment = _RIGHT

    # --- Row 34: Market Price [references Assumptions!B31] ---
    row = 34
    ws.cell(row=row, column=1, value="Market Price")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    cell = ws.cell(row=row, column=2, value="=Assumptions!$B$31")
    cell.number_format = _FMT_PRICE
    cell.alignment = _RIGHT
    cell.fill = _FACT_FILL

    # --- Row 35: % Under/Over Valued ---
    row = 35
    ws.cell(row=row, column=1, value="Implied Upside/Downside")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    formula = "=(B33-B34)/B34"
    cell = ws.cell(row=row, column=2, value=formula)
    cell.number_format = _FMT_PCT
    cell.alignment = _RIGHT

    # --- Column widths ---
    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    ws.freeze_panes = "B2"


def _write_dcf_model_v1(
    writer: pd.ExcelWriter, ctx: ValuationContext, d: dict, n: int
) -> None:
    """Write v1 EBIT-driven DCF model (no revenue data) -- also formula-based where possible."""
    a = ctx.assumptions
    yearly_ebit_at = d.get("yearly_ebit_at", [])
    yearly_fcff = d.get("yearly_fcff", [])
    yearly_pv = d.get("yearly_pv", [])

    # Create empty sheet
    empty_df = pd.DataFrame([[""] * (n + 3)])
    empty_df.to_excel(writer, sheet_name="DCF Model", index=False, header=False)
    ws = _ws(writer, "DCF Model")
    _hide_gridlines(ws)

    max_cols = n + 2  # A + n years + Terminal
    term_col = n + 2

    # Row 1: Header
    row = 1
    ws.cell(row=row, column=1, value="")
    for t in range(n):
        ws.cell(row=row, column=t + 2, value=f"Year {t+1}")
    ws.cell(row=row, column=term_col, value="Terminal")
    _style_header_row(ws, row, max_cols)

    # Row 2: EBIT(1-t) values [hardcoded since no base to formula from]
    row = 2
    ws.cell(row=row, column=1, value="EBIT(1-t)")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    for t in range(n):
        cell = ws.cell(row=row, column=t + 2, value=_safe(yearly_ebit_at[t] if t < len(yearly_ebit_at) else ""))
        if isinstance(cell.value, (int, float)):
            cell.number_format = _FMT_INT
            cell.alignment = _RIGHT
            cell.fill = _HARDCODED_FILL
            _add_comment(cell, "Hardcoded: projected from growth schedule")

    # Row 3: FCFF
    row = 3
    ws.cell(row=row, column=1, value="Free Cash Flow to Firm")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    for t in range(n):
        cell = ws.cell(row=row, column=t + 2, value=_safe(yearly_fcff[t] if t < len(yearly_fcff) else ""))
        if isinstance(cell.value, (int, float)):
            cell.number_format = _FMT_INT
            cell.alignment = _RIGHT
            cell.fill = _HARDCODED_FILL
    cell = ws.cell(row=row, column=term_col, value=_safe(d.get("terminal_value")))
    if isinstance(cell.value, (int, float)):
        cell.number_format = _FMT_INT
        cell.alignment = _RIGHT

    # Row 4: PV of FCFF
    row = 4
    ws.cell(row=row, column=1, value="PV of FCFF")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    for t in range(n):
        cell = ws.cell(row=row, column=t + 2, value=_safe(yearly_pv[t] if t < len(yearly_pv) else ""))
        if isinstance(cell.value, (int, float)):
            cell.number_format = _FMT_INT
            cell.alignment = _RIGHT
    cell = ws.cell(row=row, column=term_col, value=_safe(d.get("pv_terminal")))
    if isinstance(cell.value, (int, float)):
        cell.number_format = _FMT_INT
        cell.alignment = _RIGHT

    # Row 5: blank
    row = 6
    ws.cell(row=row, column=1, value="PV of High-Growth Phase")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    ws.cell(row=row, column=2, value=_safe(d.get("pv_high_growth")))

    row = 7
    ws.cell(row=row, column=1, value="PV of Terminal Value")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    ws.cell(row=row, column=2, value=_safe(d.get("pv_terminal")))

    row = 8
    ws.cell(row=row, column=1, value="Enterprise Value")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    ws.cell(row=row, column=2, value=_safe(d.get("enterprise_value")))

    row = 9
    ws.cell(row=row, column=1, value="Equity Value")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    ws.cell(row=row, column=2, value=_safe(d.get("equity_value")))

    row = 10
    ws.cell(row=row, column=1, value="Value per Share")
    ws.cell(row=row, column=1).font = _BOLD_FONT
    ws.cell(row=row, column=2, value=_safe(d.get("equity_value_per_share")))

    for r in range(6, 11):
        cell = ws.cell(row=r, column=2)
        if isinstance(cell.value, (int, float)):
            cell.number_format = _FMT_INT
            cell.alignment = _RIGHT

    ws.column_dimensions["A"].width = 30
    for col_idx in range(2, max_cols + 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = 15

    ws.freeze_panes = "B2"


def _write_relative_valuation(writer: pd.ExcelWriter, ctx: ValuationContext) -> None:
    """Sheet 4: Relative valuation comparison."""
    if not ctx.outputs.relative:
        return

    d = ctx.outputs.relative
    price = ctx.financials.key_stats.get("price")

    rows = []
    multiples = [
        ("P/E", "pe_value"),
        ("EV/EBITDA", "ev_ebitda_value"),
        ("P/BV", "pbv_value"),
        ("P/S", "ps_value"),
        ("Composite (Median)", "composite_value"),
    ]
    for label, key in multiples:
        val = d.get(key)
        vs_market = (val - price) / price if val and price and price > 0 else None
        rows.append([label, _safe(val), _safe(vs_market)])

    rows.append(["", "", ""])
    rows.append(["Market Price", _safe(price), ""])
    rows.append(["Discount/Premium to Composite", _safe(d.get("discount_to_composite")), ""])
    rows.append(["Methods Used", ", ".join(d.get("methods_used", [])), ""])

    if ctx.benchmarks.industry_multiples:
        rows.append(["", "", ""])
        rows.append(["INDUSTRY MULTIPLES", "Value", ""])
        for k, v in ctx.benchmarks.industry_multiples.items():
            rows.append([k, _safe(v), ""])

    df = pd.DataFrame(rows, columns=["Multiple", "Implied Value", "vs Market"])
    df.to_excel(writer, sheet_name="Relative Valuation", index=False)

    ws = _ws(writer, "Relative Valuation")
    _hide_gridlines(ws)
    _style_header_row(ws, 1, 3)

    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value or ""

        if _is_section_title(label):
            _style_section_title(ws, row_idx, 3)
            continue

        val_cell = ws.cell(row=row_idx, column=2)
        if isinstance(val_cell.value, (int, float)):
            val_cell.number_format = _FMT_PRICE
            val_cell.alignment = _RIGHT

        vs_cell = ws.cell(row=row_idx, column=3)
        if isinstance(vs_cell.value, (int, float)):
            vs_cell.number_format = _FMT_PCT
            _apply_upside_color(vs_cell)
            vs_cell.alignment = _RIGHT

    _autofit_columns(ws)
    _freeze_top_row(ws)


def _write_peer_comparison(writer: pd.ExcelWriter, ctx: ValuationContext) -> None:
    """Sheet: Peer comparison table with company vs peer median."""
    peer_data = ctx.financials.key_stats.get("peer_comparison")
    if not peer_data:
        return

    peers = peer_data.get("peers") or []
    if not peers:
        return

    rows = [
        ["PEER COMPARISON", "", "", "", "", "", ""],
    ]

    has_yahoo = any(p.get("ticker") for p in peers)

    if has_yahoo:
        rows.append(["Company", "Ticker", "P/E", "Profit Margin",
                      "Revenue Growth", "Beta", "EV/EBITDA"])
        for p in peers[:15]:
            rows.append([
                p.get("name", ""),
                p.get("ticker", ""),
                _safe(p.get("pe")),
                _safe(p.get("profit_margin")),
                _safe(p.get("revenue_growth")),
                _safe(p.get("beta")),
                _safe(p.get("ev_to_ebitda")),
            ])
    else:
        rows.append(["Company", "Revenue", "Net Income", "Total Assets",
                      "Country", "", ""])
        for p in peers[:15]:
            rows.append([
                p.get("name", ""),
                _safe(p.get("revenue")),
                _safe(p.get("net_income")),
                _safe(p.get("total_assets")),
                p.get("country", ""),
                "",
                "",
            ])

    # Peer median
    peer_median = peer_data.get("peer_median") or {}
    if peer_median:
        rows.append(["", "", "", "", "", "", ""])
        rows.append(["PEER MEDIAN", "", "", "", "", "", ""])
        for key, val in peer_median.items():
            rows.append([key, _safe(val), "", "", "", "", ""])

    # Company vs median
    vs_median = peer_data.get("company_vs_median") or {}
    if vs_median:
        rows.append(["", "", "", "", "", "", ""])
        rows.append(["COMPANY vs PEER MEDIAN", "Company", "Peer Median",
                      "Difference", "Assessment", "", ""])
        for key, comp in vs_median.items():
            rows.append([
                key,
                _safe(comp.get("company")),
                _safe(comp.get("peer_median")),
                _safe(comp.get("diff_pct")),
                comp.get("assessment", ""),
                "",
                "",
            ])

    n_cols = 7
    for r in rows:
        while len(r) < n_cols:
            r.append("")

    df = pd.DataFrame(rows, columns=[f"Col{i}" for i in range(n_cols)])
    df.to_excel(writer, sheet_name="Peer Comparison", index=False)

    ws = _ws(writer, "Peer Comparison")
    _hide_gridlines(ws)
    _style_header_row(ws, 1, n_cols)

    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value or ""
        if _is_section_title(label):
            _style_section_title(ws, row_idx, n_cols)
            continue

        # Format numeric cells
        for col_idx in range(2, n_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            if isinstance(cell.value, (int, float)):
                if abs(cell.value) <= 1:
                    cell.number_format = _FMT_PCT
                else:
                    cell.number_format = _FMT_DEC
                cell.alignment = _RIGHT

    _autofit_columns(ws)
    _freeze_top_row(ws)


def _write_sensitivity(writer: pd.ExcelWriter, ctx: ValuationContext) -> None:
    """Sheet 5: Sensitivity tables -- formatted as proper grids with WACC vs terminal growth."""
    if not ctx.outputs.sensitivity:
        return

    sens = ctx.outputs.sensitivity

    # Build base case value for highlighting
    base_case_val = None
    if ctx.outputs.dcf_fcff:
        base_case_val = ctx.outputs.dcf_fcff.get("equity_value_per_share")
    elif ctx.outputs.dcf_fcfe:
        base_case_val = ctx.outputs.dcf_fcfe.get("value_per_share")

    base_wacc = ctx.assumptions.wacc
    base_tg = ctx.assumptions.terminal_growth

    all_rows = []
    table_positions: list[dict] = []

    current_excel_row = 1

    for table_name, table_data in sens.items():
        if isinstance(table_data, dict):
            first_val = next(iter(table_data.values()), None)
            if isinstance(first_val, dict):
                # Two-way table
                col_keys = sorted(first_val.keys())
                header = [table_name] + [
                    f"{c:.4f}" if isinstance(c, float) else str(c) for c in col_keys
                ]
                all_rows.append(header)
                table_start_row = current_excel_row
                current_excel_row += 1

                row_keys = sorted(table_data.keys())
                data_start_row = current_excel_row
                for row_key in row_keys:
                    row = [f"{row_key:.4f}" if isinstance(row_key, float) else str(row_key)]
                    for ck in col_keys:
                        val = table_data[row_key].get(ck, "")
                        row.append(_safe(val))
                    all_rows.append(row)
                    current_excel_row += 1

                table_positions.append({
                    "name": table_name,
                    "header_row": table_start_row,
                    "data_start_row": data_start_row,
                    "data_end_row": current_excel_row - 1,
                    "col_keys": col_keys,
                    "row_keys": row_keys,
                    "n_cols": len(col_keys) + 1,
                    "base_wacc": base_wacc,
                    "base_tg": base_tg,
                })
            else:
                # One-way table
                all_rows.append([table_name, "Value per Share"])
                current_excel_row += 1
                for k, v in table_data.items():
                    all_rows.append([k, _safe(v)])
                    current_excel_row += 1
            # Blank spacer row
            all_rows.append(["", ""])
            current_excel_row += 1

    if not all_rows:
        return

    max_cols = max(len(r) for r in all_rows)
    for r in all_rows:
        while len(r) < max_cols:
            r.append("")

    df = pd.DataFrame(all_rows)
    df.to_excel(writer, sheet_name="Sensitivity", index=False, header=False)

    ws = _ws(writer, "Sensitivity")
    _hide_gridlines(ws)

    # Style two-way tables
    for tbl in table_positions:
        hrow = tbl["header_row"]
        ncols = tbl["n_cols"]

        _style_section_title(ws, hrow, ncols)
        ws.cell(row=hrow, column=1).font = Font(name="Calibri", bold=True, size=11)

        col_keys = tbl["col_keys"]
        row_keys = tbl["row_keys"]

        for ci, ck in enumerate(col_keys, start=2):
            cell = ws.cell(row=hrow, column=ci)
            cell.font = _BOLD_FONT
            cell.alignment = _CENTER
            if isinstance(ck, float):
                cell.value = ck
                cell.number_format = _FMT_PCT

        for ri, rk in enumerate(row_keys):
            excel_row = tbl["data_start_row"] + ri
            lbl_cell = ws.cell(row=excel_row, column=1)
            lbl_cell.font = _BOLD_FONT
            if isinstance(rk, float):
                lbl_cell.value = rk
                lbl_cell.number_format = _FMT_PCT
            lbl_cell.alignment = _RIGHT

            for ci, ck in enumerate(col_keys, start=2):
                cell = ws.cell(row=excel_row, column=ci)
                val = cell.value
                if val == "" or val is None:
                    continue
                try:
                    fval = float(val)
                    cell.value = fval
                    cell.number_format = _FMT_PRICE
                    cell.alignment = _RIGHT

                    is_base_row = (
                        base_wacc is not None
                        and isinstance(rk, float)
                        and abs(rk - base_wacc) < 1e-6
                    )
                    is_base_col = (
                        base_tg is not None
                        and isinstance(ck, float)
                        and abs(ck - base_tg) < 1e-6
                    )
                    if is_base_row and is_base_col:
                        cell.fill = PatternFill("solid", fgColor=_YELLOW_BG)
                        cell.font = Font(name="Calibri", bold=True, size=10)
                    elif base_case_val is not None:
                        if fval > base_case_val * 1.05:
                            cell.fill = PatternFill("solid", fgColor="C6EFCE")
                        elif fval < base_case_val * 0.95:
                            cell.fill = PatternFill("solid", fgColor="FFC7CE")
                except (ValueError, TypeError):
                    pass

    # Style one-way table headers
    covered_rows = set()
    for tbl in table_positions:
        covered_rows.update(range(tbl["header_row"], tbl["data_end_row"] + 1))

    for row_idx in range(1, ws.max_row + 1):
        if row_idx in covered_rows:
            continue
        cell = ws.cell(row=row_idx, column=1)
        val = cell.value
        if val and isinstance(val, str) and val.strip() and not val.strip().startswith("0."):
            if val.strip() not in ("", "Value per Share"):
                _style_section_title(ws, row_idx, 2)
            elif val.strip() == "Value per Share":
                ws.cell(row=row_idx, column=2).font = _BOLD_FONT

    _autofit_columns(ws)
    ws.freeze_panes = None


def _write_analyst_consensus(
    writer: pd.ExcelWriter, ctx: ValuationContext, ibes_data: dict | None
) -> None:
    """Sheet 6: Analyst consensus comparison -- Yahoo Finance + I/B/E/S."""
    analyst_data = ctx.financials.key_stats.get("analyst_data") or {}
    price = ctx.financials.key_stats.get("price")
    pt = analyst_data.get("price_targets") or {}
    recs = analyst_data.get("recommendations") or []
    ee = analyst_data.get("earnings_estimate") or {}

    rows = [
        ["ANALYST CONSENSUS vs OUR ESTIMATE", "", "", ""],
        ["(Consensus is for COMPARISON ONLY \u2014 not used as DCF input)", "", "", ""],
        ["", "", "", ""],
    ]

    # --- Price Targets ---
    if pt:
        rows.append(["PRICE TARGETS (YAHOO FINANCE)", "", "", ""])
        rows.append(["Mean Target", _safe(pt.get("targetMean")), "", ""])
        rows.append(["Median Target", _safe(pt.get("targetMedian")), "", ""])
        rows.append(["High Target", _safe(pt.get("targetHigh")), "", ""])
        rows.append(["Low Target", _safe(pt.get("targetLow")), "", ""])
        if pt.get("numberOfAnalysts"):
            rows.append(["Number of Analysts", pt["numberOfAnalysts"], "", ""])
        rows.append(["", "", "", ""])

    # --- Analyst Recommendations ---
    if recs:
        rows.append(["ANALYST RECOMMENDATIONS", "", "", ""])
        rows.append(["Period", "Strong Buy", "Buy", "Hold / Sell"])
        for rec in recs[:4]:
            period = rec.get("period", "")
            sb = rec.get("strongBuy", 0)
            b = rec.get("buy", 0)
            h = rec.get("hold", 0)
            s = rec.get("sell", 0) + rec.get("strongSell", 0)
            rows.append([period, sb, b, f"{h} / {s}"])
        total_latest = recs[0] if recs else {}
        total = sum(total_latest.get(k, 0) for k in ("strongBuy", "buy", "hold", "sell", "strongSell"))
        if total > 0:
            buy_pct = (total_latest.get("strongBuy", 0) + total_latest.get("buy", 0)) / total
            rows.append(["Pct Buy/Strong Buy", buy_pct, "", f"of {total} analysts"])
        rows.append(["", "", "", ""])

    # --- Earnings Estimates ---
    if ee and "avg" in ee:
        rows.append(["CONSENSUS EPS ESTIMATES", "Average", "Low", "High"])
        avg = ee.get("avg", {})
        low = ee.get("low", {})
        high = ee.get("high", {})
        growth = ee.get("growth", {})
        num = ee.get("numberOfAnalysts", {})
        for period, label in [("0q", "Current Quarter"), ("+1q", "Next Quarter"),
                               ("0y", "Current Year"), ("+1y", "Next Year")]:
            if period in avg:
                rows.append([
                    f"{label} ({period})",
                    _safe(avg.get(period)),
                    _safe(low.get(period)),
                    _safe(high.get(period)),
                ])
        rows.append(["", "", "", ""])
        rows.append(["CONSENSUS GROWTH ESTIMATES", "Growth Rate", "Num Analysts", ""])
        for period, label in [("0q", "Current Qtr"), ("+1q", "Next Qtr"),
                               ("0y", "Current Year"), ("+1y", "Next Year")]:
            if period in growth:
                rows.append([label, _safe(growth.get(period)), _safe(num.get(period)), ""])
        rows.append(["", "", "", ""])

    # --- I/B/E/S (if available) ---
    if ibes_data and "estimates" in ibes_data:
        est = ibes_data["estimates"]
        if est is not None and not est.empty:
            rows.append(["I/B/E/S CONSENSUS (WRDS)", "", "", ""])
            rows.append(["Period", "Mean EPS", "Median EPS", "Num Analysts"])
            for _, row in est.head(6).iterrows():
                rows.append([
                    str(row.get("statpers", "")),
                    _safe(row.get("meanest")),
                    _safe(row.get("medest")),
                    _safe(row.get("numest")),
                ])
            rows.append(["", "", "", ""])

    # --- Top Analysts (accuracy-ranked) ---
    top_analysts = ibes_data.get("top_analysts") if ibes_data else None
    if top_analysts is not None and not top_analysts.empty:
        rows.append(["TOP ANALYSTS (RANKED BY EPS FORECAST ACCURACY)", "", "", ""])
        rows.append(["Analyst", "Firm", "Accuracy", "Target", "Recommendation", "Num Estimates"])
        for _, row in top_analysts.iterrows():
            analyst_name = str(row.get("analyst_name") or "N/A")
            firm = str(row.get("firm") or "N/A")
            acc = row.get("accuracy_pct")
            target = row.get("target")
            rec = str(row.get("recommendation") or "N/A")
            num_est = row.get("num_estimates")
            rows.append([
                analyst_name,
                firm,
                _safe(acc / 100.0) if acc is not None else "",
                _safe(target) if target is not None else "",
                rec,
                _safe(num_est) if num_est is not None else "",
            ])
        rows.append(["", "", "", "", "", ""])

    # --- Our Estimate vs Consensus ---
    rows.append(["OUR ESTIMATE vs CONSENSUS", "Value", "vs Market", ""])
    our_dcf = None
    if ctx.outputs.dcf_fcff:
        our_dcf = ctx.outputs.dcf_fcff.get("equity_value_per_share")
        vs = (our_dcf - price) / price if price and price > 0 and our_dcf else None
        rows.append(["Our DCF (FCFF)", _safe(our_dcf), _safe(vs), "Independent fundamental"])
    if ctx.outputs.dcf_fcfe:
        our_ddm = ctx.outputs.dcf_fcfe.get("value_per_share")
        vs = (our_ddm - price) / price if price and price > 0 and our_ddm else None
        rows.append(["Our DDM", _safe(our_ddm), _safe(vs), "Independent fundamental"])
    if ctx.outputs.relative and ctx.outputs.relative.get("composite_value"):
        comp = ctx.outputs.relative["composite_value"]
        vs = (comp - price) / price if price and price > 0 else None
        rows.append(["Our Relative (composite)", _safe(comp), _safe(vs), "Industry multiples"])
    if pt.get("targetMean"):
        vs = (pt["targetMean"] - price) / price if price and price > 0 else None
        rows.append(["Analyst Mean Target", _safe(pt["targetMean"]), _safe(vs), "Yahoo Finance consensus"])
    rows.append(["Market Price", _safe(price), "", "Current"])
    rows.append(["", "", "", ""])

    # --- Why Our Valuation Differs ---
    rows.append(["WHY OUR VALUATION MAY DIFFER FROM CONSENSUS", "", "", ""])
    if our_dcf and pt.get("targetMean") and price:
        diff = our_dcf - pt["targetMean"]
        if abs(diff) > price * 0.05:
            if our_dcf < pt["targetMean"]:
                rows.append(["Our DCF is BELOW analyst targets", "", "", ""])
                rows.append(["Possible reasons:", "", "", ""])
                rows.append(["  1. We use fundamental growth (ROE x retention)", "", "", "not analyst forecasts"])
                rows.append(["  2. Our WACC may be higher (more conservative)", "", "", ""])
                rows.append(["  3. Analysts may factor in optionality/catalysts", "", "", "we don't"])
                rows.append(["  4. Our terminal growth may be lower", "", "", ""])
            else:
                rows.append(["Our DCF is ABOVE analyst targets", "", "", ""])
                rows.append(["Possible reasons:", "", "", ""])
                rows.append(["  1. Our fundamental growth may be more optimistic", "", "", ""])
                rows.append(["  2. Our WACC may be lower", "", "", ""])
                rows.append(["  3. Analysts may see risks we haven't captured", "", "", ""])
        else:
            rows.append(["Our DCF is roughly aligned with analyst targets", "", "", ""])

    # --- Growth comparison ---
    rows.append(["", "", "", ""])
    rows.append(["GROWTH: OUR ESTIMATE vs CONSENSUS", "Our", "Consensus", ""])
    if ctx.assumptions.growth_rates:
        our_g = ctx.assumptions.growth_rates[0]
        cons_g = ee.get("growth", {}).get("+1y")
        rows.append(["Year 1 Growth", _safe(our_g), _safe(cons_g), ""])
    rows.append(["Terminal Growth", _safe(ctx.assumptions.terminal_growth), "N/A", "Nominal GDP"])

    # Pad all rows to 6 columns
    n_cols = 6
    for r in rows:
        while len(r) < n_cols:
            r.append("")

    df = pd.DataFrame(rows, columns=["Item", "Value 1", "Value 2", "Notes", "Col5", "Col6"])
    df.to_excel(writer, sheet_name="Analyst Consensus", index=False)

    ws = _ws(writer, "Analyst Consensus")
    _hide_gridlines(ws)
    _style_header_row(ws, 1, n_cols)

    pct_rows = {"year 1 growth", "terminal growth", "accuracy"}
    in_top_analysts = False
    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value or ""
        label_lower = str(label).strip().lower()

        if _is_section_title(label):
            _style_section_title(ws, row_idx, n_cols)
            if "top analysts" in label_lower:
                in_top_analysts = True
            elif label_lower != "":
                in_top_analysts = False
            continue

        # Style top analysts sub-header row
        if label_lower == "analyst":
            _style_header_row(ws, row_idx, n_cols)
            in_top_analysts = True
            continue

        # Format accuracy column as percentage in top analysts block
        if in_top_analysts:
            acc_cell = ws.cell(row=row_idx, column=3)
            if isinstance(acc_cell.value, (int, float)) and abs(acc_cell.value) <= 1:
                acc_cell.number_format = _FMT_PCT
                acc_cell.alignment = _RIGHT
            tgt_cell = ws.cell(row=row_idx, column=4)
            if isinstance(tgt_cell.value, (int, float)):
                tgt_cell.number_format = _FMT_PRICE
                tgt_cell.alignment = _RIGHT

        val_cell = ws.cell(row=row_idx, column=2)
        if isinstance(val_cell.value, (int, float)):
            if label_lower in pct_rows and abs(val_cell.value) <= 3:
                val_cell.number_format = _FMT_PCT
            else:
                val_cell.number_format = _FMT_DEC
            val_cell.alignment = _RIGHT

    _autofit_columns(ws)
    _freeze_top_row(ws)


def _write_data_sources(writer: pd.ExcelWriter, ctx: ValuationContext) -> None:
    """Sheet 7: Data source transparency -- where every input came from."""
    rows = [
        ["DATA SOURCE TRANSPARENCY", "", ""],
        ["Every input is traced to its source", "", ""],
        ["", "", ""],
        ["Input", "Value", "Source"],
    ]

    stats = ctx.financials.key_stats
    a = ctx.assumptions

    rows.append(["MARKET DATA", "", ""])
    rows.append(["Price", _safe(stats.get("price")), "Yahoo Finance"])
    rows.append(["Market Cap", _safe(stats.get("market_cap")), "Yahoo Finance"])
    rows.append(["Shares Outstanding", _safe(stats.get("shares_outstanding")), "Yahoo Finance"])
    rows.append(["Beta (yfinance)", _safe(stats.get("beta")), "Yahoo Finance"])
    rows.append(["Book Value/Share", _safe(stats.get("book_value_per_share")), "Yahoo Finance"])
    rows.append(["Dividend/Share", _safe(stats.get("dividend_per_share")), "Yahoo Finance"])

    rows.append(["", "", ""])
    rows.append(["RISK INPUTS", "", ""])
    rows.append(["Risk-Free Rate", _safe(a.risk_free_rate), "Govt bond yield"])
    rows.append(["ERP", _safe(a.erp), "Damodaran histimpl.xls"])
    rows.append(["CRP", _safe(a.country_risk_premium), "Damodaran ctryprem.xlsx"])
    rows.append(["Industry Unlevered Beta", _safe(ctx.benchmarks.industry_unlevered_beta), f"Damodaran betas ({ctx.company.region})"])
    rows.append(["Beta (re-levered)", _safe(a.beta), "Computed: Bu*(1+(1-t)*D/E)"])
    rows.append(["Cost of Equity", _safe(a.cost_of_equity), "Computed: CAPM"])
    rows.append(["Cost of Debt", _safe(a.cost_of_debt), "Computed: Rf + synthetic spread"])
    rows.append(["WACC", _safe(a.wacc), "Computed"])
    rows.append(["Tax Rate", _safe(a.tax_rate), "Computed from financials"])

    rows.append(["", "", ""])
    rows.append(["GROWTH INPUTS", "", ""])
    rows.append(["Terminal Growth", _safe(a.terminal_growth), "Nominal GDP growth"])
    if a.growth_rates:
        rows.append(["High Growth Rate", _safe(a.growth_rates[0]), "See Assumptions sheet"])

    rows.append(["", "", ""])
    rows.append(["FINANCIAL DATA", "", ""])
    if ctx.financials.income_statement is not None:
        rows.append(["Income Statement", f"{ctx.financials.income_statement.shape[0]} years", "Yahoo Finance"])
    if ctx.financials.balance_sheet is not None:
        rows.append(["Balance Sheet", f"{ctx.financials.balance_sheet.shape[0]} years", "Yahoo Finance"])
    if ctx.financials.cash_flow is not None:
        rows.append(["Cash Flow", f"{ctx.financials.cash_flow.shape[0]} years", "Yahoo Finance"])

    rows.append(["", "", ""])
    rows.append(["INDUSTRY BENCHMARKS", "", ""])
    rows.append(["Damodaran Industry", ctx.company.damodaran_industry or "N/A", "Fuzzy match"])
    rows.append(["Industry WACC", _safe(ctx.benchmarks.industry_wacc), f"Damodaran wacc ({ctx.company.region})"])

    pct_labels_ds = {"risk-free rate", "erp", "crp", "cost of equity", "cost of debt",
                     "wacc", "tax rate", "terminal growth", "industry wacc",
                     "beta (re-levered)", "industry unlevered beta"}

    df = pd.DataFrame(rows, columns=["Input", "Value", "Source"])
    df.to_excel(writer, sheet_name="Data Sources", index=False)

    ws = _ws(writer, "Data Sources")
    _hide_gridlines(ws)
    _style_header_row(ws, 1, 3)

    for row_idx in range(2, ws.max_row + 1):
        label = ws.cell(row=row_idx, column=1).value or ""

        if _is_section_title(label):
            _style_section_title(ws, row_idx, 3)
            continue

        val_cell = ws.cell(row=row_idx, column=2)
        if isinstance(val_cell.value, (int, float)):
            label_lower = str(label).lower()
            if label_lower in pct_labels_ds and abs(val_cell.value) <= 3:
                val_cell.number_format = _FMT_PCT
            elif label_lower in ("price", "book value/share", "dividend/share"):
                val_cell.number_format = _FMT_PRICE
            elif label_lower in ("market cap", "shares outstanding"):
                val_cell.number_format = _FMT_INT
            else:
                val_cell.number_format = _FMT_DEC
            val_cell.alignment = _RIGHT

    _autofit_columns(ws)
    _freeze_top_row(ws)


_IS_COL_ORDER: dict[str, list[str]] = {
    "Income Statement": [
        "Total Revenue", "Cost Of Revenue", "Gross Profit",
        "Operating Expense", "Selling General And Administration", "Research And Development",
        "Operating Income", "Interest Expense", "Other Income Expense",
        "Pretax Income", "Tax Provision", "Net Income",
        "EBITDA", "EBIT", "Basic EPS", "Diluted EPS",
    ],
    "Balance Sheet": [
        "Cash And Cash Equivalents", "Current Assets", "Total Assets",
        "Current Liabilities", "Long Term Debt", "Total Debt", "Total Liabilities",
        "Total Stockholders Equity", "Share Issued",
    ],
    "Cash Flow": [
        "Operating Cash Flow", "Capital Expenditure", "Free Cash Flow",
        "Depreciation And Amortization", "Stock Based Compensation",
        "Change In Working Capital", "Investing Cash Flow", "Financing Cash Flow",
    ],
}


def _reorder_financial_rows(df: pd.DataFrame, sheet_name: str) -> pd.DataFrame:
    """Reorder rows (line items) of a financial statement DataFrame into standard order."""
    canonical = _IS_COL_ORDER.get(sheet_name, [])
    if not canonical:
        return df

    existing = {str(idx).lower(): idx for idx in df.index}

    ordered_idx: list = []
    seen: set = set()
    for name in canonical:
        key = name.lower()
        if key in existing:
            orig = existing[key]
            ordered_idx.append(orig)
            seen.add(orig)

    for idx in df.index:
        if idx not in seen:
            ordered_idx.append(idx)

    return df.loc[ordered_idx]


def _write_financials(writer: pd.ExcelWriter, ctx: ValuationContext) -> None:
    """Sheet 8: Raw financial statements -- rows=line items, columns=fiscal years."""
    sheets = [
        ("income_statement", "Income Statement"),
        ("balance_sheet", "Balance Sheet"),
        ("cash_flow", "Cash Flow"),
    ]

    for attr, sheet_name in sheets:
        df_raw: pd.DataFrame | None = getattr(ctx.financials, attr, None)
        if df_raw is None:
            continue

        df = df_raw.copy()

        index_is_dates = _looks_like_dates(df.index)
        cols_are_dates = _looks_like_dates(df.columns)

        if index_is_dates and not cols_are_dates:
            df = df.T
        elif cols_are_dates:
            pass

        try:
            df = df.sort_index(axis=1, ascending=False)
        except TypeError:
            pass

        df = _reorder_financial_rows(df, sheet_name)
        df.to_excel(writer, sheet_name=sheet_name)

        ws = _ws(writer, sheet_name)
        _hide_gridlines(ws)
        n_cols = df.shape[1] + 1

        _style_header_row(ws, 1, n_cols)

        for row_idx in range(2, ws.max_row + 1):
            label_cell = ws.cell(row=row_idx, column=1)
            label = label_cell.value or ""
            label_cell.font = _NORMAL_FONT
            label_cell.alignment = _LEFT

            for col_idx in range(2, n_cols + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if isinstance(cell.value, (int, float)):
                    label_lower = str(label).lower()
                    if any(kw in label_lower for kw in ("margin", "rate", "ratio", "yield", "growth")):
                        if abs(cell.value) <= 5:
                            cell.number_format = _FMT_PCT
                        else:
                            cell.number_format = _FMT_INT
                    else:
                        cell.number_format = _FMT_INT
                    cell.alignment = _RIGHT

                    if cell.value < 0:
                        cell.font = Font(name="Calibri", size=10, color=_RED)

                    # All financial statement data is from source (green)
                    cell.fill = _FACT_FILL

        ws.column_dimensions["A"].width = 35
        for col_idx in range(2, n_cols + 1):
            ws.column_dimensions[get_column_letter(col_idx)].width = 15

        _freeze_top_row(ws)


# ---------------------------------------------------------------------------
# Internal helpers for financials orientation detection
# ---------------------------------------------------------------------------

def _looks_like_dates(index) -> bool:
    """Return True if the index/columns appear to contain date-like values."""
    if len(index) == 0:
        return False
    sample = index[0]
    if isinstance(sample, pd.Timestamp):
        return True
    try:
        from datetime import date as _date, datetime as _datetime
        if isinstance(sample, (_date, _datetime)):
            return True
    except ImportError:
        pass
    if isinstance(sample, str):
        s = sample.strip()
        if len(s) == 4 and s.isdigit():
            return True
        if len(s) >= 7 and s[:4].isdigit() and s[4] in ("-", "/"):
            return True
    return False
