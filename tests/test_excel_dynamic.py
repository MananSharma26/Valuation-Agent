"""Tests verifying the Excel workbook is fully dynamic with cross-sheet references.

The Assumptions sheet has a fixed layout and the DCF Model sheet references it,
so changing any assumption on the Assumptions sheet automatically updates the DCF.
"""

from __future__ import annotations

import tempfile
from pathlib import Path

import openpyxl
import pytest

from valuation.context import ValuationContext
from valuation.reports.excel_writer import generate_excel


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------


def _make_v2_ctx() -> ValuationContext:
    """Context with revenue-based (v2) DCF output."""
    ctx = ValuationContext("TEST")
    ctx.company.name = "Test Corp"
    ctx.company.sector = "Technology"
    ctx.company.classification = "mature"
    ctx.company.region = "US"
    ctx.company.damodaran_industry = "Software"

    ctx.financials.key_stats = {
        "price": 100.0,
        "shares_outstanding": 1_000_000,
        "market_cap": 100_000_000,
        "beta": 1.1,
        "book_value_per_share": 25.0,
    }

    ctx.assumptions.risk_free_rate = 0.04
    ctx.assumptions.erp = 0.05
    ctx.assumptions.country_risk_premium = 0.0
    ctx.assumptions.beta = 1.1
    ctx.assumptions.cost_of_equity = 0.095
    ctx.assumptions.cost_of_debt = 0.05
    ctx.assumptions.wacc = 0.085
    ctx.assumptions.tax_rate = 0.21
    ctx.assumptions.terminal_growth = 0.025
    ctx.assumptions.projection_years = 10
    ctx.assumptions.growth_rates = [0.15, 0.14, 0.13, 0.12, 0.11, 0.10, 0.09, 0.08, 0.07, 0.06]

    n = 10
    base_rev = 10_000_000
    yearly_rev = []
    r = base_rev
    for g in ctx.assumptions.growth_rates:
        r = r * (1 + g)
        yearly_rev.append(r)

    margin = 0.20
    yearly_ebit = [rev * margin for rev in yearly_rev]
    yearly_ebit_at = [ebit * (1 - 0.21) for ebit in yearly_ebit]
    yearly_reinv = [
        (yearly_rev[t] - (base_rev if t == 0 else yearly_rev[t - 1])) / 2.0
        for t in range(n)
    ]
    yearly_fcff = [yearly_ebit_at[t] - yearly_reinv[t] for t in range(n)]

    wacc = 0.085
    yearly_pv = []
    cum = 1.0
    for t in range(n):
        cum *= (1 + wacc)
        yearly_pv.append(yearly_fcff[t] / cum)

    ctx.outputs.dcf_fcff = {
        "enterprise_value": 50_000_000,
        "equity_value": 45_000_000,
        "equity_value_per_share": 45.0,
        "pv_high_growth": sum(yearly_pv),
        "pv_terminal": 30_000_000,
        "terminal_value": 60_000_000,
        "terminal_fcff": yearly_ebit_at[-1] * 1.025 * (1 - 0.025 / 0.20),
        "yearly_revenue": yearly_rev,
        "yearly_ebit": yearly_ebit,
        "yearly_ebit_at": yearly_ebit_at,
        "yearly_reinvestment": yearly_reinv,
        "yearly_fcff": yearly_fcff,
        "yearly_pv": yearly_pv,
        "yearly_ic": [],
        "yearly_roic": [],
    }

    ctx.outputs.sensitivity = {}
    ctx.confidence.composite = 0.75

    return ctx


@pytest.fixture
def workbook(tmp_path):
    """Generate the workbook and return the loaded openpyxl Workbook."""
    ctx = _make_v2_ctx()
    path = generate_excel(ctx, ibes_data=None, output_path=tmp_path / "test.xlsx")
    return openpyxl.load_workbook(str(path))


# ---------------------------------------------------------------------------
# Assumptions sheet: fixed layout
# ---------------------------------------------------------------------------


class TestAssumptionsFixedLayout:
    def test_risk_free_rate_at_b4(self, workbook):
        ws = workbook["Assumptions"]
        assert ws.cell(4, 2).value == 0.04

    def test_beta_at_b5(self, workbook):
        ws = workbook["Assumptions"]
        assert ws.cell(5, 2).value == 1.1

    def test_erp_at_b6(self, workbook):
        ws = workbook["Assumptions"]
        assert ws.cell(6, 2).value == 0.05

    def test_lambda_at_b7(self, workbook):
        ws = workbook["Assumptions"]
        # Lambda should be a number (default 1.0 when CRP=0)
        assert isinstance(ws.cell(7, 2).value, (int, float))

    def test_crp_at_b8(self, workbook):
        ws = workbook["Assumptions"]
        assert ws.cell(8, 2).value == 0.0

    def test_cost_of_equity_formula_at_b9(self, workbook):
        ws = workbook["Assumptions"]
        val = ws.cell(9, 2).value
        assert isinstance(val, str) and "B4" in val and "B5" in val and "B6" in val

    def test_wacc_formula_at_b14(self, workbook):
        ws = workbook["Assumptions"]
        val = ws.cell(14, 2).value
        assert isinstance(val, str) and "B9" in val and "B13" in val

    def test_high_growth_at_b17(self, workbook):
        ws = workbook["Assumptions"]
        assert ws.cell(17, 2).value == 0.15

    def test_terminal_growth_at_b18(self, workbook):
        ws = workbook["Assumptions"]
        assert ws.cell(18, 2).value == 0.025

    def test_base_revenue_at_b26(self, workbook):
        ws = workbook["Assumptions"]
        assert isinstance(ws.cell(26, 2).value, (int, float))
        assert ws.cell(26, 2).value > 0

    def test_shares_at_b30(self, workbook):
        ws = workbook["Assumptions"]
        assert ws.cell(30, 2).value == 1_000_000

    def test_price_at_b31(self, workbook):
        ws = workbook["Assumptions"]
        assert ws.cell(31, 2).value == 100.0


# ---------------------------------------------------------------------------
# DCF Model sheet: cross-sheet references
# ---------------------------------------------------------------------------


class TestDCFModelCrossSheetReferences:
    def test_growth_year1_references_assumptions(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(3, 3).value  # C3 = Year 1 growth
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$17" in val

    def test_growth_terminal_references_assumptions(self, workbook):
        ws = workbook["DCF Model"]
        # Terminal col = 10 + 3 = 13
        val = ws.cell(3, 13).value
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$18" in val

    def test_base_revenue_references_assumptions(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(4, 2).value  # B4 = base revenue
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$26" in val

    def test_revenue_year1_formula(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(4, 3).value  # C4 = Year 1 revenue
        assert isinstance(val, str)
        assert "B4" in val and "C3" in val

    def test_margin_year1_references_assumptions(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(5, 3).value  # C5 = Year 1 margin
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$19" in val  # current margin
        assert "B$20" in val  # target margin

    def test_ebit_formula(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(6, 3).value  # C6 = Revenue * Margin
        assert isinstance(val, str)
        assert "C4" in val and "C5" in val

    def test_tax_year1_references_assumptions(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(9, 3).value  # C9 = Tax rate
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$11" in val

    def test_s2c_references_assumptions(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(11, 3).value  # C11 = S2C
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$21" in val

    def test_wacc_year1_references_assumptions(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(16, 3).value  # C16 = WACC Year 1
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$14" in val

    def test_cash_references_assumptions(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(28, 2).value  # B28 = Cash
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$28" in val

    def test_debt_references_assumptions(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(29, 2).value  # B29 = Debt
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$29" in val

    def test_shares_references_assumptions(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(32, 2).value  # B32 = Shares
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$30" in val

    def test_price_references_assumptions(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(34, 2).value  # B34 = Price
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$31" in val

    def test_terminal_fcff_formula(self, workbook):
        ws = workbook["DCF Model"]
        # Terminal FCFF in col 13 (n+3), row 13
        val = ws.cell(13, 13).value
        assert isinstance(val, str)
        assert "Assumptions" in val
        assert "B$18" in val  # terminal growth
        assert "B$22" in val  # stable ROC

    def test_value_per_share_formula(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(33, 2).value
        assert val == "=B31/B32"

    def test_upside_formula(self, workbook):
        ws = workbook["DCF Model"]
        val = ws.cell(35, 2).value
        assert val == "=(B33-B34)/B34"


# ---------------------------------------------------------------------------
# No gridlines
# ---------------------------------------------------------------------------


class TestNoGridlines:
    def test_assumptions_no_gridlines(self, workbook):
        ws = workbook["Assumptions"]
        assert ws.sheet_view.showGridLines is False

    def test_dcf_model_no_gridlines(self, workbook):
        ws = workbook["DCF Model"]
        assert ws.sheet_view.showGridLines is False
